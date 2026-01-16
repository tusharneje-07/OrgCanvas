"""
Excel Parser Module
Handles reading and writing organizational data from/to Excel files.

Excel Schema:
| employee_id | name | title | department | manager_id | avatar_url | color |
|-------------|------|-------|------------|------------|------------|-------|
| 1           | John | CEO   | Executive  | None       | /avatar.png| #1a73e8 |

- employee_id: Unique identifier for each employee (required)
- name: Employee's full name (required)
- title: Job title/position (required)
- department: Department name (required)
- manager_id: ID of the manager (None/empty for root node)
- avatar_url: Optional URL/path to avatar image
- color: Optional hex color for the node
"""

import os
from typing import Dict, List, Optional, Any
from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException
import logging

logger = logging.getLogger(__name__)

# Excel column mapping
EXCEL_COLUMNS = {
    'employee_id': 'A',
    'name': 'B',
    'title': 'C',
    'department': 'D',
    'manager_id': 'E',
    'avatar_url': 'F',
    'color': 'G',
    'whatsapp': 'H'
}

# Default department colors
DEFAULT_DEPARTMENT_COLORS = {
    'Executive': '#1a73e8',
    'Dean': '#1a73e8',
    "Dean's Office": '#1a73e8',
    'Community': '#1a73e8',
    'Graduate Studies': '#1a73e8',
    'Programs': '#1a73e8',
    'Academic Affairs': '#4caf50',
    'Development': '#4caf50',
    'Facilities': '#4caf50',
    'Finance': '#4caf50',
    'Exhibitions': '#4caf50',
    'Auxiliary Staff': '#f57c00',
    'HR': '#9c27b0',
    'IT': '#00bcd4',
    'Marketing': '#e91e63',
    'Sales': '#ff5722',
    'Operations': '#795548',
    'Legal': '#607d8b',
    'default': '#757575'
}


class ExcelParserError(Exception):
    """Custom exception for Excel parsing errors."""
    pass


class ExcelParser:
    """
    Parses and writes organizational hierarchy data from/to Excel files.
    
    Attributes:
        file_path (str): Path to the Excel file
        data (List[Dict]): Parsed employee data
    """
    
    def __init__(self, file_path: str):
        """
        Initialize the Excel parser.
        
        Args:
            file_path: Path to the Excel file
        """
        self.file_path = file_path
        self.data: List[Dict[str, Any]] = []
        self._workbook: Optional[Workbook] = None
    
    def parse(self) -> List[Dict[str, Any]]:
        """
        Parse the Excel file and return a list of employee dictionaries.
        
        Returns:
            List of employee dictionaries with all required fields
            
        Raises:
            ExcelParserError: If file cannot be read or has invalid format
        """
        if not os.path.exists(self.file_path):
            raise ExcelParserError(f"Excel file not found: {self.file_path}")
        
        try:
            self._workbook = load_workbook(filename=self.file_path, read_only=True)
            sheet = self._workbook.active
            
            if sheet is None:
                raise ExcelParserError("No active sheet found in Excel file")
            
            # Get headers from first row
            headers = []
            for cell in sheet[1]:
                headers.append(str(cell.value).lower().strip() if cell.value else '')
            
            # Validate required columns
            required_columns = ['employee_id', 'name', 'title', 'department']
            for col in required_columns:
                if col not in headers:
                    raise ExcelParserError(f"Missing required column: {col}")
            
            # Parse data rows
            self.data = []
            for row_num, row in enumerate(sheet.iter_rows(min_row=2), start=2):
                row_data = {}
                for idx, cell in enumerate(row):
                    if idx < len(headers) and headers[idx]:
                        row_data[headers[idx]] = cell.value
                
                # Skip empty rows
                if not row_data.get('employee_id') or not row_data.get('name'):
                    continue
                
                # Process and validate employee data
                employee = self._process_row(row_data, row_num)
                if employee:
                    self.data.append(employee)
            
            self._workbook.close()
            logger.info(f"Successfully parsed {len(self.data)} employees from Excel")
            return self.data
            
        except InvalidFileException as e:
            raise ExcelParserError(f"Invalid Excel file format: {e}")
        except Exception as e:
            raise ExcelParserError(f"Error parsing Excel file: {e}")
    
    def _process_row(self, row_data: Dict, row_num: int) -> Optional[Dict[str, Any]]:
        """
        Process a single row of data into a standardized employee dictionary.
        
        Args:
            row_data: Raw row data from Excel
            row_num: Row number for error reporting
            
        Returns:
            Processed employee dictionary or None if invalid
        """
        try:
            employee_id = str(row_data.get('employee_id', '')).strip()
            if not employee_id:
                logger.warning(f"Row {row_num}: Missing employee_id, skipping")
                return None
            
            name = str(row_data.get('name', '')).strip()
            if not name:
                logger.warning(f"Row {row_num}: Missing name, skipping")
                return None
            
            title = str(row_data.get('title', '')).strip()
            department = str(row_data.get('department', '')).strip()
            
            # Handle manager_id - can be None, empty, or "None" string
            manager_id = row_data.get('manager_id')
            if manager_id is None or str(manager_id).strip().lower() in ['', 'none', 'null']:
                manager_id = None
            else:
                manager_id = str(manager_id).strip()
            
            # Get optional fields
            avatar_url = row_data.get('avatar_url', '')
            if avatar_url:
                avatar_url = str(avatar_url).strip()
            
            # Get color - use department default if not specified
            color = row_data.get('color', '')
            if color:
                color = str(color).strip()
            if not color:
                color = DEFAULT_DEPARTMENT_COLORS.get(
                    department, 
                    DEFAULT_DEPARTMENT_COLORS['default']
                )
            
            # Get WhatsApp number
            whatsapp = row_data.get('whatsapp', '')
            if whatsapp:
                # Clean the number - remove spaces, dashes, etc.
                whatsapp = str(whatsapp).strip().replace(' ', '').replace('-', '').replace('(', '').replace(')', '')
                # Remove leading + if present for storage, we'll add it back for display
                if whatsapp.startswith('+'):
                    whatsapp = whatsapp[1:]
            
            return {
                'id': employee_id,
                'name': name,
                'title': title,
                'department': department,
                'manager_id': manager_id,
                'avatar_url': avatar_url if avatar_url else None,
                'color': color,
                'whatsapp': whatsapp if whatsapp else None
            }
            
        except Exception as e:
            logger.error(f"Row {row_num}: Error processing row - {e}")
            return None
    
    def save(self, data: List[Dict[str, Any]], file_path: Optional[str] = None) -> bool:
        """
        Save employee data back to Excel file.
        
        Args:
            data: List of employee dictionaries to save
            file_path: Optional alternative file path to save to
            
        Returns:
            True if save successful, False otherwise
            
        Raises:
            ExcelParserError: If save fails
        """
        save_path = file_path or self.file_path
        
        try:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Organization"
            
            # Write headers
            headers = ['employee_id', 'name', 'title', 'department', 'manager_id', 'avatar_url', 'color', 'whatsapp']
            for col_num, header in enumerate(headers, 1):
                sheet.cell(row=1, column=col_num, value=header)
            
            # Write data rows
            for row_num, employee in enumerate(data, 2):
                sheet.cell(row=row_num, column=1, value=employee.get('id', ''))
                sheet.cell(row=row_num, column=2, value=employee.get('name', ''))
                sheet.cell(row=row_num, column=3, value=employee.get('title', ''))
                sheet.cell(row=row_num, column=4, value=employee.get('department', ''))
                sheet.cell(row=row_num, column=5, value=employee.get('manager_id') or '')
                sheet.cell(row=row_num, column=6, value=employee.get('avatar_url') or '')
                sheet.cell(row=row_num, column=7, value=employee.get('color', ''))
                sheet.cell(row=row_num, column=8, value=employee.get('whatsapp') or '')
            
            # Adjust column widths
            column_widths = {'A': 12, 'B': 20, 'C': 25, 'D': 20, 'E': 12, 'F': 30, 'G': 10, 'H': 15}
            for col, width in column_widths.items():
                sheet.column_dimensions[col].width = width
            
            workbook.save(save_path)
            logger.info(f"Successfully saved {len(data)} employees to Excel")
            return True
            
        except PermissionError:
            raise ExcelParserError(
                f"Cannot save to {save_path}: File is locked or permission denied. "
                "Please close the file if it's open in another application."
            )
        except Exception as e:
            raise ExcelParserError(f"Error saving Excel file: {e}")
    
    def validate_data(self, data: List[Dict[str, Any]]) -> List[str]:
        """
        Validate employee data for consistency and integrity.
        
        Args:
            data: List of employee dictionaries to validate
            
        Returns:
            List of validation error messages (empty if valid)
        """
        errors = []
        employee_ids = set()
        
        for employee in data:
            emp_id = employee.get('id')
            
            # Check for duplicate IDs
            if emp_id in employee_ids:
                errors.append(f"Duplicate employee_id: {emp_id}")
            employee_ids.add(emp_id)
            
            # Check for required fields
            if not employee.get('name'):
                errors.append(f"Employee {emp_id}: Missing name")
            if not employee.get('title'):
                errors.append(f"Employee {emp_id}: Missing title")
            if not employee.get('department'):
                errors.append(f"Employee {emp_id}: Missing department")
        
        # Validate manager references
        for employee in data:
            manager_id = employee.get('manager_id')
            if manager_id and manager_id not in employee_ids:
                errors.append(
                    f"Employee {employee.get('id')}: Invalid manager_id '{manager_id}' - "
                    "manager does not exist"
                )
        
        return errors


def create_sample_excel(file_path: str) -> bool:
    """
    Create a sample Excel file with demo organizational data.
    
    Args:
        file_path: Path where the sample file should be created
        
    Returns:
        True if creation successful
    """
    sample_data = [
        {'id': '1', 'name': 'Peter Murphy', 'title': 'Dean', 'department': 'Executive', 'manager_id': None, 'color': '#1a73e8', 'whatsapp': '1234567890'},
        {'id': '2', 'name': 'Ronald Cox', 'title': 'VP Operations', 'department': 'Operations', 'manager_id': '1', 'color': '#f57c00', 'whatsapp': None},
        {'id': '3', 'name': 'Marvin Lee', 'title': 'VP Academics', 'department': 'Academics', 'manager_id': '1', 'color': '#4caf50', 'whatsapp': '9876543210'},
        {'id': '4', 'name': 'Kate Williams', 'title': 'Director', 'department': 'Operations', 'manager_id': '2', 'color': '#f57c00', 'whatsapp': None},
        {'id': '5', 'name': 'Holly Greene', 'title': 'Manager', 'department': 'Operations', 'manager_id': '2', 'color': '#f57c00', 'whatsapp': None},
        {'id': '6', 'name': 'Silvia Lewis', 'title': 'Director', 'department': 'Academics', 'manager_id': '3', 'color': '#4caf50', 'whatsapp': None},
        {'id': '7', 'name': 'Lydia Chance', 'title': 'Manager', 'department': 'Academics', 'manager_id': '3', 'color': '#4caf50', 'whatsapp': None},
    ]
    
    parser = ExcelParser(file_path)
    return parser.save(sample_data, file_path)
