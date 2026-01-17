"""
Organization Chart Maker - Flask Application
A web application for viewing and editing organizational hierarchies.

Routes:
- GET  /                : Main application page
- POST /upload_excel    : Upload new Excel file
- GET  /get_org_data    : Get organizational data as JSON
- POST /update_node     : Update a single node
- POST /add_node        : Add a new node
- POST /delete_node     : Delete a node
- POST /save_excel      : Save changes to Excel file
- POST /export          : Export chart as PNG or PDF
"""

import os
import io
import base64
import json
import logging
from datetime import datetime
from typing import Optional

from flask import (
    Flask, render_template, request, jsonify, 
    send_file, redirect, url_for
)
from werkzeug.utils import secure_filename

from utils.excel_parser import ExcelParser, ExcelParserError, create_sample_excel
from utils.tree_builder import TreeBuilder, TreeBuilderError

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Flask application configuration
app = Flask(__name__)
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'dev-secret-key-change-in-production')
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max upload
app.config['UPLOAD_FOLDER'] = os.path.join(os.path.dirname(__file__), 'uploads')
app.config['DATA_FOLDER'] = os.path.join(os.path.dirname(__file__), 'data')

# Allowed file extensions
ALLOWED_EXTENSIONS = {'xlsx', 'xls'}

# Global state for organizational data
class OrgDataStore:
    """
    Singleton store for organizational data.
    Maintains in-memory state and tracks the source file.
    """
    def __init__(self):
        self.tree_builder: Optional[TreeBuilder] = None
        self.source_file: Optional[str] = None
        self.last_modified: Optional[datetime] = None
        self.has_unsaved_changes: bool = False
    
    def load_from_excel(self, file_path: str) -> dict:
        """Load data from Excel file and build tree."""
        parser = ExcelParser(file_path)
        employees = parser.parse()
        
        # Validate data
        errors = parser.validate_data(employees)
        if errors:
            logger.warning(f"Validation warnings: {errors}")
        
        # Build tree
        self.tree_builder = TreeBuilder(employees)
        self.tree_builder.build()
        self.source_file = file_path
        self.last_modified = datetime.now()
        self.has_unsaved_changes = False
        
        return self.tree_builder.to_dict()
    
    def get_data(self) -> Optional[dict]:
        """Get current tree data as dictionary."""
        if self.tree_builder:
            return self.tree_builder.to_dict()
        return None
    
    def save_to_excel(self, file_path: Optional[str] = None) -> bool:
        """Save current data to Excel file."""
        if not self.tree_builder:
            raise Exception("No data loaded")
        
        save_path = file_path or self.source_file
        if not save_path:
            raise Exception("No file path specified")
        
        flat_list = self.tree_builder.to_flat_list()
        parser = ExcelParser(save_path)
        result = parser.save(flat_list, save_path)
        
        if result:
            self.has_unsaved_changes = False
            self.last_modified = datetime.now()
        
        return result
    
    def mark_changed(self):
        """Mark data as having unsaved changes."""
        self.has_unsaved_changes = True
        self.last_modified = datetime.now()


# Initialize data store
org_store = OrgDataStore()


def allowed_file(filename: str) -> bool:
    """Check if file extension is allowed."""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def ensure_directories():
    """Create necessary directories if they don't exist."""
    for folder in [app.config['UPLOAD_FOLDER'], app.config['DATA_FOLDER']]:
        if not os.path.exists(folder):
            os.makedirs(folder)


def load_default_data():
    """Load default/sample data on startup."""
    ensure_directories()
    
    default_file = os.path.join(app.config['DATA_FOLDER'], 'organization.xlsx')
    
    # Create sample file if it doesn't exist
    if not os.path.exists(default_file):
        logger.info("Creating sample organization data...")
        create_sample_excel(default_file)
    
    # Load the data
    try:
        org_store.load_from_excel(default_file)
        logger.info(f"Loaded organization data from {default_file}")
    except Exception as e:
        logger.error(f"Failed to load default data: {e}")


# Load data on startup
with app.app_context():
    load_default_data()


# =============================================================================
# ROUTES
# =============================================================================

@app.route('/')
def index():
    """Render the main application page."""
    return render_template('index.html')


@app.route('/upload_excel', methods=['POST'])
def upload_excel():
    """
    Handle Excel file upload.
    
    Request: multipart/form-data with 'file' field
    Response: JSON with success status and organization data
    """
    try:
        # Check if file is in request
        if 'file' not in request.files:
            return jsonify({
                'success': False,
                'error': 'No file provided'
            }), 400
        
        file = request.files['file']
        
        if file.filename == '':
            return jsonify({
                'success': False,
                'error': 'No file selected'
            }), 400
        
        if not allowed_file(file.filename):
            return jsonify({
                'success': False,
                'error': 'Invalid file type. Please upload an Excel file (.xlsx or .xls)'
            }), 400
        
        # Save uploaded file
        ensure_directories()
        filename = secure_filename(file.filename)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        saved_filename = f"{timestamp}_{filename}"
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], saved_filename)
        file.save(file_path)
        
        # Load the data
        data = org_store.load_from_excel(file_path)
        
        return jsonify({
            'success': True,
            'message': f'Successfully loaded {data["total_employees"]} employees',
            'data': data
        })
        
    except ExcelParserError as e:
        logger.error(f"Excel parsing error: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 400
    except TreeBuilderError as e:
        logger.error(f"Tree building error: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 400
    except Exception as e:
        logger.exception(f"Unexpected error during upload: {e}")
        return jsonify({
            'success': False,
            'error': 'An unexpected error occurred while processing the file'
        }), 500


@app.route('/get_org_data', methods=['GET'])
def get_org_data():
    """
    Get current organizational data as JSON.
    
    Response: JSON with organization tree structure
    """
    try:
        data = org_store.get_data()
        
        if data is None:
            return jsonify({
                'success': False,
                'error': 'No organization data loaded'
            }), 404
        
        return jsonify({
            'success': True,
            'data': data,
            'has_unsaved_changes': org_store.has_unsaved_changes,
            'source_file': os.path.basename(org_store.source_file) if org_store.source_file else None
        })
        
    except Exception as e:
        logger.exception(f"Error getting org data: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/update_node', methods=['POST'])
def update_node():
    """
    Update a single node in the organization.
    
    Request JSON:
    {
        "id": "employee_id",
        "name": "New Name",
        "title": "New Title",
        "department": "New Dept",
        "manager_id": "new_manager_id",
        "color": "#hex"
    }
    
    Response: JSON with updated node data
    """
    try:
        data = request.get_json()
        
        if not data or 'id' not in data:
            return jsonify({
                'success': False,
                'error': 'Node ID is required'
            }), 400
        
        if not org_store.tree_builder:
            return jsonify({
                'success': False,
                'error': 'No organization data loaded'
            }), 404
        
        node_id = str(data['id'])
        updates = {k: v for k, v in data.items() if k != 'id'}
        
        # Update the node
        updated_node = org_store.tree_builder.update_node(node_id, updates)
        
        if not updated_node:
            return jsonify({
                'success': False,
                'error': f'Node with ID {node_id} not found'
            }), 404
        
        org_store.mark_changed()
        
        return jsonify({
            'success': True,
            'message': 'Node updated successfully',
            'node': updated_node.to_dict(),
            'data': org_store.tree_builder.to_dict()
        })
        
    except TreeBuilderError as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 400
    except Exception as e:
        logger.exception(f"Error updating node: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/add_node', methods=['POST'])
def add_node():
    """
    Add a new node to the organization.
    
    Request JSON:
    {
        "name": "Employee Name",
        "title": "Job Title",
        "department": "Department",
        "manager_id": "manager_id" (optional)
    }
    
    Response: JSON with new node data
    """
    try:
        data = request.get_json()
        
        if not data:
            return jsonify({
                'success': False,
                'error': 'No data provided'
            }), 400
        
        required_fields = ['name', 'title', 'department']
        for field in required_fields:
            if not data.get(field):
                return jsonify({
                    'success': False,
                    'error': f'{field} is required'
                }), 400
        
        if not org_store.tree_builder:
            return jsonify({
                'success': False,
                'error': 'No organization data loaded'
            }), 404
        
        # Generate new ID
        new_id = org_store.tree_builder.generate_new_id()
        
        # Prepare employee data
        employee_data = {
            'id': new_id,
            'name': data['name'],
            'title': data['title'],
            'department': data['department'],
            'manager_id': data.get('manager_id'),
            'avatar_url': data.get('avatar_url'),
            'color': data.get('color', '#757575')
        }
        
        # Add the node
        new_node = org_store.tree_builder.add_node(employee_data)
        org_store.mark_changed()
        
        return jsonify({
            'success': True,
            'message': 'Node added successfully',
            'node': new_node.to_dict(),
            'data': org_store.tree_builder.to_dict()
        })
        
    except TreeBuilderError as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 400
    except Exception as e:
        logger.exception(f"Error adding node: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/delete_node', methods=['POST'])
def delete_node():
    """
    Delete a node from the organization.
    
    Request JSON:
    {
        "id": "employee_id",
        "reassign_to": "other_employee_id" (optional)
    }
    
    Response: JSON with success status
    """
    try:
        data = request.get_json()
        
        if not data or 'id' not in data:
            return jsonify({
                'success': False,
                'error': 'Node ID is required'
            }), 400
        
        if not org_store.tree_builder:
            return jsonify({
                'success': False,
                'error': 'No organization data loaded'
            }), 404
        
        node_id = str(data['id'])
        reassign_to = data.get('reassign_to')
        
        # Delete the node
        org_store.tree_builder.delete_node(node_id, reassign_to)
        org_store.mark_changed()
        
        return jsonify({
            'success': True,
            'message': 'Node deleted successfully',
            'data': org_store.tree_builder.to_dict()
        })
        
    except TreeBuilderError as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 400
    except Exception as e:
        logger.exception(f"Error deleting node: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/save_excel', methods=['POST'])
def save_excel():
    """
    Save current organization data to Excel file.
    
    Request JSON (optional):
    {
        "filename": "custom_filename.xlsx"
    }
    
    Response: JSON with success status
    """
    try:
        data = request.get_json() or {}
        
        if not org_store.tree_builder:
            return jsonify({
                'success': False,
                'error': 'No organization data loaded'
            }), 404
        
        # Determine save path
        if 'filename' in data:
            ensure_directories()
            filename = secure_filename(data['filename'])
            if not filename.endswith('.xlsx'):
                filename += '.xlsx'
            save_path = os.path.join(app.config['DATA_FOLDER'], filename)
        else:
            save_path = org_store.source_file
        
        if not save_path:
            return jsonify({
                'success': False,
                'error': 'No file path specified'
            }), 400
        
        # Save to Excel
        org_store.save_to_excel(save_path)
        
        return jsonify({
            'success': True,
            'message': 'Changes saved successfully',
            'filename': os.path.basename(save_path)
        })
        
    except ExcelParserError as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 400
    except Exception as e:
        logger.exception(f"Error saving Excel: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/export', methods=['POST'])
def export_chart():
    """
    Export the organization chart as PNG or PDF.
    
    Request JSON:
    {
        "format": "png" or "pdf",
        "image_data": "base64 encoded image data",
        "pdf_options": {  // Optional, for PDF only
            "pageSize": "auto" | "a4" | "a3" | "a2" | "a1" | "a0" | "letter" | "legal" | "tabloid",
            "orientation": "landscape" | "portrait",
            "margin": 10  // in mm
        }
    }
    
    Response: File download
    """
    try:
        data = request.get_json()
        
        if not data:
            return jsonify({
                'success': False,
                'error': 'No data provided'
            }), 400
        
        export_format = data.get('format', 'png').lower()
        image_data = data.get('image_data', '')
        pdf_options = data.get('pdf_options', {})
        
        if not image_data:
            return jsonify({
                'success': False,
                'error': 'No image data provided'
            }), 400
        
        # Remove data URL prefix if present
        if ',' in image_data:
            image_data = image_data.split(',')[1]
        
        # Decode base64 image
        image_bytes = base64.b64decode(image_data)
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        if export_format == 'pdf':
            # Convert to PDF using reportlab
            from reportlab.lib.pagesizes import A4, A3, A2, A1, A0, LETTER, LEGAL, TABLOID, landscape, portrait
            from reportlab.pdfgen import canvas
            from reportlab.lib.utils import ImageReader
            from reportlab.lib.units import mm
            from PIL import Image
            
            # Load image
            img = Image.open(io.BytesIO(image_bytes))
            img_width, img_height = img.size
            
            # Get PDF options
            page_size_name = pdf_options.get('pageSize', 'auto')
            orientation = pdf_options.get('orientation', 'landscape')
            margin_mm = pdf_options.get('margin', 10)
            margin = margin_mm * mm
            
            # Page size mapping
            PAGE_SIZES = {
                'a4': A4,
                'a3': A3,
                'a2': A2,
                'a1': A1,
                'a0': A0,
                'letter': LETTER,
                'legal': LEGAL,
                'tabloid': TABLOID
            }
            
            # Create PDF buffer
            pdf_buffer = io.BytesIO()
            
            if page_size_name == 'auto':
                # Auto-fit: Create page size based on image dimensions with margin
                # Convert pixels to points (72 points per inch, assume 150 DPI for the image)
                dpi = 150  # Assumed DPI of the high-quality export
                points_per_pixel = 72 / dpi
                
                page_width = (img_width * points_per_pixel) + (2 * margin)
                page_height = (img_height * points_per_pixel) + (2 * margin)
                page_size = (page_width, page_height)
                
                c = canvas.Canvas(pdf_buffer, pagesize=page_size)
                
                # Draw image at full size with margin
                scaled_width = img_width * points_per_pixel
                scaled_height = img_height * points_per_pixel
                x = margin
                y = margin
                
            else:
                # Use specified page size
                base_page_size = PAGE_SIZES.get(page_size_name, A4)
                
                # Apply orientation
                if orientation == 'landscape':
                    page_size = landscape(base_page_size)
                else:
                    page_size = portrait(base_page_size)
                
                c = canvas.Canvas(pdf_buffer, pagesize=page_size)
                page_width, page_height = page_size
                
                # Calculate scaling to fit page with margins
                available_width = page_width - 2 * margin
                available_height = page_height - 2 * margin
                
                scale_x = available_width / img_width
                scale_y = available_height / img_height
                scale = min(scale_x, scale_y)
                
                # Calculate centered position
                scaled_width = img_width * scale
                scaled_height = img_height * scale
                x = (page_width - scaled_width) / 2
                y = (page_height - scaled_height) / 2
            
            # Draw image with high quality settings
            img_reader = ImageReader(img)
            c.drawImage(img_reader, x, y, width=scaled_width, height=scaled_height, 
                       preserveAspectRatio=True, mask='auto')
            
            c.save()
            pdf_buffer.seek(0)
            
            return send_file(
                pdf_buffer,
                mimetype='application/pdf',
                as_attachment=True,
                download_name=f'org_chart_{timestamp}.pdf'
            )
        else:
            # Return PNG
            img_buffer = io.BytesIO(image_bytes)
            img_buffer.seek(0)
            
            return send_file(
                img_buffer,
                mimetype='image/png',
                as_attachment=True,
                download_name=f'org_chart_{timestamp}.png'
            )
            
    except Exception as e:
        logger.exception(f"Error exporting chart: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/download_template', methods=['GET'])
def download_template():
    """Download a template Excel file with the correct schema."""
    try:
        ensure_directories()
        template_path = os.path.join(app.config['DATA_FOLDER'], 'template.xlsx')
        
        # Create template with headers and sample row
        from openpyxl import Workbook
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Organization"
        
        # Headers
        headers = ['employee_id', 'name', 'title', 'department', 'manager_id', 'avatar_url', 'color']
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Sample row
        sample = ['1', 'John Doe', 'CEO', 'Executive', '', '', '#1a73e8']
        for col, value in enumerate(sample, 1):
            ws.cell(row=2, column=col, value=value)
        
        # Column widths
        widths = {'A': 12, 'B': 20, 'C': 25, 'D': 20, 'E': 12, 'F': 30, 'G': 10}
        for col, width in widths.items():
            ws.column_dimensions[col].width = width
        
        wb.save(template_path)
        
        return send_file(
            template_path,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='org_chart_template.xlsx'
        )
        
    except Exception as e:
        logger.exception(f"Error creating template: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


# Error handlers
@app.errorhandler(404)
def not_found(error):
    return jsonify({
        'success': False,
        'error': 'Resource not found'
    }), 404


@app.errorhandler(500)
def internal_error(error):
    return jsonify({
        'success': False,
        'error': 'Internal server error'
    }), 500


if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)
